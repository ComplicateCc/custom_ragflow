#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
RAG 测试脚本
用于循环测试RAG接口并收集检索结果
"""

import os
import sys
import json
import time
import pandas as pd
import numpy as np
from datetime import datetime
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# 导入RAGFlow SDK
from ragflow_sdk import RAGFlow

# 配置参数
API_KEY = "ragflow-E3ZDc1ZWUwMGEwZDExZjBhODkzMDI0Mm"  # 替换为您的API密钥
HOST_ADDRESS = "http://localhost:9380"  # 替换为您的RAGFlow服务地址
# TEST_DATASET_NAME = "客服中台-bge_embedding"  # 替换为您要测试的现有数据集名称
# TEST_DATASET_NAME = "客服中台-Tongyi"  # 替换为您要测试的现有数据集名称
TEST_DATASET_NAME = "客服中台-仅保留Question录入测试"

# 测试数据文件路径 - 使用绝对路径
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEST_DATA_FILE = os.path.join(SCRIPT_DIR, "test_data", "现有节点非相似问的玩家问话.xlsx")

# 创建结果目录
RESULTS_DIR = os.path.join(SCRIPT_DIR, "rag_test_results")
os.makedirs(RESULTS_DIR, exist_ok=True)

def load_test_data(excel_file):
    """从Excel文件加载测试数据"""
    print(f"正在加载测试数据: {excel_file}")
    try:
        df = pd.read_excel(excel_file, header=0)
        # 读取第一列作为问题，第二列作为标准答案
        questions = df.iloc[:, 0].tolist()
        answers = df.iloc[:, 1].tolist()
        
        # 过滤空值
        valid_data = [(q, a) for q, a in zip(questions, answers) if isinstance(q, str) and isinstance(a, str)]
        
        if not valid_data:
            raise ValueError("Excel文件中没有有效的问题和答案数据")
        
        questions, answers = zip(*valid_data)
        print(f"成功加载 {len(questions)} 对问题和标准答案")
        
        return list(questions), list(answers)
    except Exception as e:
        print(f"加载测试数据时出错: {str(e)}")
        return [], []

def get_dataset_documents(dataset):
    """获取数据集中的所有文档"""
    print(f"正在获取数据集 '{dataset.name}' 中的文档...")
    
    # 获取第一页文档
    page = 1
    page_size = 100
    all_documents = []
    
    while True:
        documents_page = dataset.list_documents(page=page, page_size=page_size)
        if not documents_page:
            break
            
        all_documents.extend(documents_page)
        print(f"已获取 {len(all_documents)} 个文档")
        
        if len(documents_page) < page_size:
            break
            
        page += 1
    
    print(f"数据集中共有 {len(all_documents)} 个文档")
    return all_documents

def test_rag_queries(rag_instance, dataset, questions, answers=None, documents=None):
    """测试RAG查询并收集结果"""
    results = []
    dataset_ids = [dataset.id]
    document_ids = [doc.id for doc in documents] if documents else []
    
    print(f"开始测试 {len(questions)} 个问题...")
    for i, question in enumerate(tqdm(questions)):
        # 直接调用HTTP API而不是SDK的retrieve方法
        start_time = time.time()
        response = rag_instance.post('/retrieval', json={
            "dataset_ids": dataset_ids,
            "documents": document_ids,
            "question": question,
            "page": 1,
            "page_size": 10,
            "similarity_threshold": 0.6,
            "vector_similarity_weight": 0.8,
            "top_k": 1024,
            "rerank_id": None,
            "keyword": False
        })
        end_time = time.time()
        
        response_data = response.json()
        
        if response_data.get("code") != 0:
            print(f"API调用错误: {response_data.get('message', '未知错误')}")
            chunks = []
        else:
            chunks = response_data.get("data", {}).get("chunks", [])
        
        # 收集结果
        query_result = {
            "question": question,
            "standard_answer": answers[i] if answers and i < len(answers) else "",
            "retrieval_time": end_time - start_time,
            "retrieval_count": len(chunks),
            "chunks": []
        }
        
        # 收集检索到的chunks详情
        for chunk in chunks:
            chunk_data = {
                "content": chunk.get("content", ""),
                "document_id": chunk.get("document_id", ""),
                "document_name": chunk.get("document_keyword", chunk.get("document_name", "未知文档")),
                "similarity": chunk.get("similarity", 0.0),
                "vector_similarity": chunk.get("vector_similarity", 0.0),
                "term_similarity": chunk.get("term_similarity", 0.0),
            }
            query_result["chunks"].append(chunk_data)
        
        results.append(query_result)
    
    return results

def save_results(results, output_dir):
    """保存测试结果"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    result_file = os.path.join(output_dir, f"rag_test_results_{timestamp}.json")
    
    with open(result_file, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    
    print(f"测试结果已保存到: {result_file}")
    
    # 生成简单的统计报告
    report_file = os.path.join(output_dir, f"rag_test_report_{timestamp}.txt")
    with open(report_file, "w", encoding="utf-8") as f:
        f.write("RAG测试报告\n")
        f.write("=" * 50 + "\n\n")
        
        # 总体统计
        f.write(f"测试时间: {timestamp}\n")
        f.write(f"测试问题数: {len(results)}\n\n")
        
        # 每个问题的统计
        for i, result in enumerate(results):
            f.write(f"问题 {i+1}: {result['question']}\n")
            f.write(f"检索时间: {result['retrieval_time']:.4f} 秒\n")
            f.write(f"检索结果数: {result['retrieval_count']}\n")
            
            # 最高相似度的前3个结果
            if result['chunks']:
                f.write("前3个最相关的结果:\n")
                sorted_chunks = sorted(result['chunks'], key=lambda x: x['similarity'], reverse=True)[:3]
                for j, chunk in enumerate(sorted_chunks):
                    f.write(f"  {j+1}. 相似度: {chunk['similarity']:.4f}, 文档: {chunk['document_name']}\n")
                    f.write(f"     内容: {chunk['content'][:100]}...\n")
            
            f.write("\n" + "-" * 40 + "\n\n")
    
    print(f"测试报告已生成: {report_file}")
    return result_file, report_file

def extract_answer_from_content(content):
    """从内容中提取Answer后面的部分"""
    if "Answer: " in content:
        return content.split("Answer: ", 1)[1].strip()
    return content.strip()

def analyze_results(results, answers):
    """分析测试结果并与标准答案对比"""
    total_questions = len(results)
    correct_count = 0
    comparison_results = []
    
    for result in results:
        question = result['question']
        standard_answer = result['standard_answer']
        
        # 获取相似度最高的Top5 chunks作为RAG答案候选
        rag_answer = ""
        similarity = 0.0
        if result['chunks']:
            # 排序并获取Top5
            top_chunks = sorted(result['chunks'], key=lambda x: x['similarity'], reverse=True)[:5]
            
            # 统计相同Answer内容出现的频率
            answer_frequency = {}
            for chunk in top_chunks:
                # 从内容中提取Answer部分
                content = chunk['content'].strip()
                extracted_answer = extract_answer_from_content(content)
                
                if extracted_answer not in answer_frequency:
                    answer_frequency[extracted_answer] = {
                        'count': 0,
                        'similarity': 0,
                        'content': content,  # 保存原始内容
                        'answer': extracted_answer  # 保存提取出的答案
                    }
                answer_frequency[extracted_answer]['count'] += 1
                answer_frequency[extracted_answer]['similarity'] = max(
                    answer_frequency[extracted_answer]['similarity'], 
                    chunk['similarity']
                )
            
            # 找出频率最高的答案
            if answer_frequency:
                # 先按频率排序，如果频率相同，则按相似度排序
                best_answer = max(answer_frequency.values(), key=lambda x: (x['count'], x['similarity']))
                rag_answer = best_answer['content']  # 使用原始内容作为回答
                extracted_answer = best_answer['answer']  # 提取的答案部分
                similarity = best_answer['similarity']
        
        # 判断RAG答案是否包含标准答案的关键内容
        is_correct = False
        if rag_answer and standard_answer:
            # 提取RAG答案中的Answer部分进行比较
            extracted_answer = extract_answer_from_content(rag_answer)
            # 判断标准答案是否包含在提取的答案中，或者提取的答案是否包含标准答案
            is_correct = standard_answer in extracted_answer or extracted_answer in standard_answer
        
        if is_correct:
            correct_count += 1
        
        comparison_results.append({
            'question': question,
            'rag_answer': rag_answer,
            'extracted_answer': extract_answer_from_content(rag_answer) if rag_answer else "",
            'similarity': similarity,
            'standard_answer': standard_answer,
            'is_correct': is_correct,
            'top_chunks': top_chunks if result['chunks'] else []
        })
    
    accuracy = correct_count / total_questions if total_questions > 0 else 0
    
    print("\n============ RAG测试分析 ============")
    print(f"测试问题总数: {total_questions}")
    print(f"正确回答数: {correct_count}")
    print(f"准确率: {accuracy:.2%}")
    print("====================================\n")
    
    return comparison_results, accuracy

def export_to_excel(comparison_results, accuracy, output_dir):
    """将对比结果导出为Excel格式"""
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = os.path.join(output_dir, f"rag_comparison_{timestamp}.xlsx")
    
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = "RAG测试结果对比"
    
    # 设置表头
    headers = ["问题", "RAG答案", "提取答案", "相似度", "正确答案", "是否正确", "Top5候选答案"]
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num).value = header
    
    # 设置单元格样式
    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    # 填充数据
    for row_num, result in enumerate(comparison_results, 2):
        ws.cell(row=row_num, column=1).value = result['question']
        ws.cell(row=row_num, column=2).value = result['rag_answer']
        ws.cell(row=row_num, column=3).value = result['extracted_answer']
        ws.cell(row=row_num, column=4).value = result['similarity']
        ws.cell(row=row_num, column=5).value = result['standard_answer']
        ws.cell(row=row_num, column=6).value = "是" if result['is_correct'] else "否"
        
        # 添加Top5候选答案信息
        top_chunks_info = ""
        if 'top_chunks' in result and result['top_chunks']:
            for i, chunk in enumerate(result['top_chunks']):
                extracted = extract_answer_from_content(chunk['content'])
                top_chunks_info += f"{i+1}. 相似度: {chunk['similarity']:.4f}\n   原始: {chunk['content'][:50]}...\n   提取: {extracted[:50]}...\n"
        ws.cell(row=row_num, column=7).value = top_chunks_info
        
        # 设置是否正确单元格的背景色
        if result['is_correct']:
            ws.cell(row=row_num, column=6).fill = green_fill
        else:
            ws.cell(row=row_num, column=6).fill = red_fill
    
    # 添加总结行
    summary_row = len(comparison_results) + 3
    ws.cell(row=summary_row, column=1).value = "总结"
    ws.cell(row=summary_row, column=2).value = f"总问题数: {len(comparison_results)}"
    ws.cell(row=summary_row, column=3).value = f"正确数: {sum(1 for r in comparison_results if r['is_correct'])}"
    ws.cell(row=summary_row, column=4).value = f"准确率: {accuracy:.2%}"
    
    # 保存Excel文件
    wb.save(excel_file)
    print(f"对比结果已导出到Excel文件: {excel_file}")
    
    return excel_file

def main():
    try:
        print("启动RAG测试...")
        
        # 加载测试数据
        test_questions, test_answers = load_test_data(TEST_DATA_FILE)
        if not test_questions:
            raise Exception("无法加载测试数据，请检查Excel文件格式")
        
        # 初始化RAGFlow客户端
        rag = RAGFlow(API_KEY, HOST_ADDRESS)
        print(f"成功连接到RAGFlow服务: {HOST_ADDRESS}")
        
        # 获取现有数据集
        dataset = rag.get_dataset(name=TEST_DATASET_NAME)
        if not dataset:
            raise Exception(f"未找到数据集: {TEST_DATASET_NAME}")
        print(f"成功获取数据集: {dataset.name} (ID: {dataset.id})")
        
        # 获取数据集中的文档
        documents = get_dataset_documents(dataset)
        if not documents:
            print("警告: 数据集中没有文档，将在所有数据集范围内测试")
        else:
            print(f"将使用数据集中的 {len(documents)} 个文档进行测试")
        
        # 运行测试查询
        results = test_rag_queries(rag, dataset, test_questions, test_answers, documents)
        
        # 保存测试结果
        result_file, report_file = save_results(results, RESULTS_DIR)
        
        # 分析结果并与标准答案对比
        comparison_results, accuracy = analyze_results(results, test_answers)
        
        # 导出对比结果到Excel
        excel_file = export_to_excel(comparison_results, accuracy, RESULTS_DIR)
        
        print(f"测试完成! 结果保存在: {RESULTS_DIR}")
        
    except Exception as e:
        print(f"测试过程中发生错误: {str(e)}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == "__main__":
    sys.exit(main())